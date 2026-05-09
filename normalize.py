"""
normalize.py — приводит файл в порядок без переклассификации.

Что делает:
1. Нормализует названия типов (Продажный → Продающий, ОБРАЗОВАТЕЛЬНЫЙ → Образовательный)
2. Обновляет обоснование типа на стандартное из кодификатора
3. Пересчитывает комбинации формат+тип
4. Применяет ручные исправления из manual_review если передан
5. Пересчитывает аналитику

НЕ меняет: коды типов, коды форматов, тексты постов, медиа-данные

Запуск:
    python normalize.py \\
      --input result_final_v2_run20260506_1052.xlsx \\
      --manual-review result_final_v2_run20260506_1052_manual_review.xlsx \\
      --output result_final_v3.xlsx
"""

import argparse
import logging
import sys
from datetime import datetime
from pathlib import Path

import openpyxl

logging.basicConfig(
    level=logging.INFO,
    format="[%(asctime)s] %(levelname)s — %(message)s",
    datefmt="%Y-%m-%d %H:%M:%S",
)
logger = logging.getLogger("normalize")

TYPE_NAMES = {
    1: "Информационный",
    2: "Образовательный",
    3: "Анонсовый",
    4: "Продающий",
    5: "Вовлекающий",
    6: "Развлекательный",
    7: "Имиджево-вдохновляющий",
}

TYPE_REASONS = {
    1: "Пост передаёт актуальные сведения или новости без побуждения к немедленному действию.",
    2: "Пост обучает аудиторию — содержит инструкцию, совет, разбор или пошаговое объяснение.",
    3: "Пост анонсирует предстоящее событие — создаёт ожидание и подталкивает к планированию.",
    4: "Пост стимулирует покупку — содержит явное коммерческое предложение, цену или призыв к заказу.",
    5: "Пост вовлекает аудиторию в диалог — содержит вопрос, опрос, конкурс или призыв к реакции.",
    6: "Пост создаёт лёгкий эмоциональный фон — юмор, мем или ситуативный контент без утилитарной ценности.",
    7: "Пост транслирует философию и ценности бренда — эстетика, вдохновение, backstage без продающего посыла.",
}

# Все варианты кривых названий → правильный код
NAME_TO_CODE = {
    # Информационный
    "информационный": 1, "информационное": 1, "информация": 1,
    "informational": 1, "information": 1,
    # Образовательный
    "образовательный": 2, "образовательное": 2, "образование": 2,
    "образовывать": 2, "образователь": 2, "обучающий": 2,
    "educational": 2, "education": 2,
    # Анонсовый
    "анонсовый": 3, "анонсовое": 3, "анонс": 3, "анонсирующий": 3,
    "анонсированный": 3, "анонсирован": 3, "анонсовой": 3,
    "announcement": 3,
    # Продающий
    "продающий": 4, "продающее": 4, "продажный": 4, "продажное": 4,
    "продажа": 4, "продавающий": 4, "продавая": 4, "продвигающий": 4,
    "продаючий": 4, "продаоучий": 4, "продавщий": 4,
    "sales": 4, "selling": 4,
    # Вовлекающий
    "вовлекающий": 5, "вовлекающее": 5, "вовлечение": 5,
    "engaging": 5, "engagement": 5,
    # Развлекательный
    "развлекательный": 6, "развлекательное": 6, "развлечение": 6,
    "разлукательный": 6, "разлюкательный": 6,
    "entertainment": 6,
    # Имиджево-вдохновляющий
    "имиджево-вдохновляющий": 7, "имиджево вдохновляющий": 7,
    "имиджевый": 7, "вдохновляющий": 7, "имиджево-вдохновляющее": 7,
    "имиджево-вдохновляющий": 7,
    "brand image": 7, "inspirational": 7,
    # Не определено
    "не определено": None, "undefined": None,
}


def normalize_type_name(name, code) -> tuple:
    """Возвращает (нормализованный код, нормализованное название)."""
    if code and str(code).strip().isdigit():
        c = int(str(code).strip())
        if c in TYPE_NAMES:
            return c, TYPE_NAMES[c]

    if name:
        name_lower = str(name).strip().lower()
        if name_lower in NAME_TO_CODE:
            c = NAME_TO_CODE[name_lower]
            if c:
                return c, TYPE_NAMES[c]

    return code, name


def parse_args():
    parser = argparse.ArgumentParser()
    parser.add_argument("--input",         required=True)
    parser.add_argument("--manual-review", default="", dest="manual_review")
    parser.add_argument("--output",        default="result_normalized.xlsx")
    return parser.parse_args()


def main():
    args = parse_args()
    logger.info("=" * 60)
    logger.info("Нормализация файла результатов")
    logger.info("=" * 60)

    if not Path(args.input).exists():
        logger.error(f"Файл не найден: {args.input}")
        sys.exit(1)

    # Загружаем основной файл
    wb = openpyxl.load_workbook(args.input)
    ws = wb["posts_raw"] if "posts_raw" in wb.sheetnames else wb.active

    rows = list(ws.iter_rows(values_only=False))
    headers = [str(cell.value).strip() if cell.value else f"col_{i}"
               for i, cell in enumerate(rows[0])]

    def col(name, aliases=[]):
        for n in [name] + aliases:
            if n in headers:
                return headers.index(n)
        return None

    url_idx      = col("Ссылка на пост")
    type_idx     = col("Тип", ["Название типа"])
    code_idx     = col("Код типа")
    reason_idx   = col("Обоснование типа", ["Обоснование"])
    combo_idx    = col("Комбинация")
    combo_c_idx  = col("Код комбо")
    format_idx   = col("Формат", ["Название формата"])
    format_c_idx = col("Код формата")
    basis_idx    = col("Основа классификации")
    review_idx   = col("Требует проверки")

    logger.info(f"Загружено строк: {len(rows)-1}")

    # Загружаем ручные исправления
    manual_fixes = {}
    if args.manual_review and Path(args.manual_review).exists():
        wb_mr = openpyxl.load_workbook(args.manual_review)
        ws_mr = wb_mr.active
        mr_rows = list(ws_mr.iter_rows(values_only=True))
        mr_headers = [str(h) for h in mr_rows[0]]
        mr_url_idx  = mr_headers.index("Ссылка на пост") if "Ссылка на пост" in mr_headers else None
        mr_fix_idx  = mr_headers.index("Исправленный код типа") if "Исправленный код типа" in mr_headers else None

        if mr_url_idx is not None and mr_fix_idx is not None:
            for row in mr_rows[1:]:
                url = row[mr_url_idx]
                fix = row[mr_fix_idx]
                if url and fix and str(fix).strip():
                    try:
                        manual_fixes[str(url)] = int(str(fix).strip())
                    except ValueError:
                        pass
        logger.info(f"Ручных исправлений: {len(manual_fixes)}")

    # Нормализуем данные
    stats = {"normalized": 0, "manual_applied": 0, "reason_updated": 0}

    for row in rows[1:]:
        vals = {h: row[i].value for i, h in enumerate(headers) if i < len(row)}

        url       = str(vals.get("Ссылка на пост", "") or "")
        type_name = str(vals.get("Тип", "") or vals.get("Название типа", "") or "")
        type_code = vals.get("Код типа")
        format_name = str(vals.get("Формат", "") or vals.get("Название формата", "") or "")
        format_code = vals.get("Код формата")

        # Применяем ручное исправление (нормализуем URL для сравнения)
        url_clean = url.strip().rstrip("/")
        matched_url = None
        for fix_url in manual_fixes:
            if fix_url.strip().rstrip("/") == url_clean:
                matched_url = fix_url
                break
        if matched_url or url in manual_fixes:
            url = matched_url or url
        if url in manual_fixes:
            new_code = manual_fixes[url]
            type_code = new_code
            type_name = TYPE_NAMES[new_code]
            if code_idx is not None: row[code_idx].value = new_code
            if type_idx is not None: row[type_idx].value = type_name
            if reason_idx is not None:
                row[reason_idx].value = TYPE_REASONS.get(new_code, "Тип исправлен вручную.")
            if basis_idx is not None: row[basis_idx].value = "manual"
            if review_idx is not None: row[review_idx].value = "Нет"
            stats["manual_applied"] += 1

        else:
            # Нормализуем название типа
            new_code, new_name = normalize_type_name(type_name, type_code)

            if new_code != type_code or new_name != type_name:
                if code_idx is not None: row[code_idx].value = new_code
                if type_idx is not None: row[type_idx].value = new_name
                type_code = new_code
                type_name = new_name
                stats["normalized"] += 1

            # Обновляем обоснование если оно кривое (короткое или на англ)
            current_reason = str(vals.get("Обоснование типа", "") or
                                  vals.get("Обоснование", "") or "")
            reason_is_bad = (
                len(current_reason) < 10 or
                sum(1 for c in current_reason if c.isascii() and c.isalpha()) /
                max(len(current_reason), 1) > 0.3
            )
            if reason_is_bad and type_code and int(str(type_code)) in TYPE_REASONS:
                if reason_idx is not None:
                    row[reason_idx].value = TYPE_REASONS[int(str(type_code))]
                    stats["reason_updated"] += 1

        # Пересчитываем комбинацию
        if type_code and format_code:
            combo_code = f"{format_code}.{type_code}"
            combo_name = f"{format_name} + {TYPE_NAMES.get(int(str(type_code)), str(type_code))}"
            if combo_c_idx is not None: row[combo_c_idx].value = combo_code
            if combo_idx is not None: row[combo_idx].value = combo_name

    logger.info(f"Нормализовано названий: {stats['normalized']}")
    logger.info(f"Применено ручных правок: {stats['manual_applied']}")
    logger.info(f"Обновлено обоснований: {stats['reason_updated']}")

    # Сохраняем
    p = Path(args.output)
    run_str = datetime.now().strftime("%Y%m%d_%H%M")
    output_path = str(p.parent / f"{p.stem}_run{run_str}{p.suffix}")
    wb.save(output_path)
    logger.info(f"Сохранено: {output_path}")

    # Пересчитываем аналитику
    logger.info("Пересчитываем аналитику...")
    try:
        import pandas as pd
        from report_builder import build_report

        df = pd.read_excel(output_path, sheet_name="posts_raw")
        posts = df.to_dict("records")

        # Нормализуем ключи
        key_map = {
            "URL канала": "channel_url", "Username": "channel_username",
            "ID поста": "post_id", "Ссылка на пост": "post_url",
            "Дата публикации": "post_date", "Месяц": "month",
            "Сегмент (EN)": "segment", "Сегмент": "segment_name",
            "Текст поста": "text", "Код формата": "format_code",
            "Формат": "format_name", "Код типа": "type_code",
            "Тип": "type_name", "Уверенность": "type_confidence",
            "Обоснование типа": "type_reason", "Код комбо": "combo_code",
            "Комбинация": "combo_name", "Основа классификации": "analysis_basis",
            "Требует проверки": "manual_review_required",
        }
        norm_posts = []
        for p_raw in posts:
            p_norm = {}
            for k, v in p_raw.items():
                p_norm[key_map.get(str(k), str(k))] = v
            norm_posts.append(p_norm)

        final_path = output_path.replace(".xlsx", "_final.xlsx")
        build_report(norm_posts, final_path)
        logger.info(f"Финальный отчёт: {final_path}")
        print(f"\n✅ Готово!")
        print(f"   Финальный отчёт: {final_path}")
        print(f"   Нормализовано названий: {stats['normalized']}")
        print(f"   Применено ручных правок: {stats['manual_applied']}")
        print(f"   Обновлено обоснований: {stats['reason_updated']}")
    except Exception as e:
        logger.warning(f"Пересчёт аналитики не удался: {e}")
        print(f"\n✅ Нормализация завершена!")
        print(f"   Файл: {output_path}")
        print(f"   Нормализовано: {stats['normalized']}, правок: {stats['manual_applied']}")


if __name__ == "__main__":
    main()
