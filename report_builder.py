"""
report_builder.py — формирование Excel-отчёта с 14 листами.

Листы:
  1.  posts_raw
  2.  format_stats_total
  3.  type_stats_total
  4.  combo_stats_total
  5.  format_stats_by_segment
  6.  type_stats_by_segment
  7.  combo_stats_by_segment
  8.  combo_pivot_total
  9.  combo_pivot_fashion
  10. combo_pivot_beauty
  11. segment_comparison
  12. stable_combinations
  13. rare_combinations
  14. method_notes + summary_text
"""

from __future__ import annotations

import csv
import logging
from datetime import datetime
from pathlib import Path

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from analytics import Analytics

logger = logging.getLogger("tg_analyzer")

# ---------------------------------------------------------------------------
# Цвета и стили
# ---------------------------------------------------------------------------
BLUE     = "2D5B9E"
YELLOW   = "FFF3CD"
GREEN    = "D6F0D6"
ALT_ROW  = "F0F4FB"
ORANGE   = "FFE0B2"
PINK     = "FCE4EC"

def _thin():
    s = Side(style="thin")
    return Border(left=s, right=s, top=s, bottom=s)

def _hdr(cell, color=BLUE):
    try:
        cell.fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
        cell.font = Font(color="FFFFFF", bold=True, size=10)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = _thin()
    except Exception:
        pass

def _data(cell, row_idx: int, alt=True):
    try:
        fill_color = ALT_ROW if (alt and row_idx % 2 == 0) else "FFFFFF"
        cell.fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")
        cell.alignment = Alignment(vertical="top", wrap_text=False)
        cell.border = _thin()
    except Exception:
        pass

def _auto_width(ws, min_w=8, max_w=50):
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            try:
                if cell.value:
                    max_len = max(max_len, len(str(cell.value)))
            except Exception:
                pass
        ws.column_dimensions[col_letter].width = min(max(max_len + 2, min_w), max_w)

def _safe_val(val):
    """Защищает значение ячейки от интерпретации как формулы Excel."""
    if isinstance(val, str) and val and val[0] in ("=", "+", "-", "@", "|"):
        return "'" + val
    return val


def _write_table(ws, headers: list[str], rows: list[dict], freeze="A2"):
    """Записывает таблицу на лист с заголовком и данными."""
    ws.append(headers)
    for cell in ws[1]:
        _hdr(cell)
    ws.row_dimensions[1].height = 30

    for row_idx, row in enumerate(rows, start=2):
        values = [_safe_val(row.get(h)) for h in headers]
        ws.append(values)
        for cell in ws[row_idx]:
            _data(cell, row_idx)

    if freeze:
        ws.freeze_panes = freeze
    _auto_width(ws)


# ---------------------------------------------------------------------------
# POSTS_RAW — колонки итоговой таблицы постов
# ---------------------------------------------------------------------------

POST_COLUMNS = [
    "channel_url", "channel_username", "post_id", "post_url",
    "post_date", "month", "segment", "segment_name",
    "text",
    "has_text", "has_photo", "has_video", "has_video_note",
    "has_voice", "has_audio", "has_poll", "has_gif", "has_sticker", "has_file",
    "media_count",
    "format_code", "format_name",
    "type_code", "type_name",
    "type_confidence", "type_reason", "type_evidence",
    "combo_code", "combo_name",
    "image_ocr_text", "image_description",
    "video_frame_ocr_text", "video_description",
    "audio_transcript", "voice_transcript",
    "analysis_basis", "llm_used", "manual_review_required",
    "classification_note",
    "views", "forwards", "replies", "reactions",
    "random_seed", "sample_order",
]

POST_HEADERS_RU = {
    "channel_url": "URL канала",
    "channel_username": "Username",
    "post_id": "ID поста",
    "post_url": "Ссылка на пост",
    "post_date": "Дата публикации",
    "month": "Месяц",
    "segment": "Сегмент (EN)",
    "segment_name": "Сегмент",
    "text": "Текст поста",
    "has_text": "Есть текст",
    "has_photo": "Фото",
    "has_video": "Видео",
    "has_video_note": "Кружок",
    "has_voice": "Голосовое",
    "has_audio": "Аудио",
    "has_poll": "Опрос",
    "has_gif": "GIF",
    "has_sticker": "Стикер",
    "has_file": "Файл",
    "media_count": "Кол-во медиа",
    "format_code": "Код формата",
    "format_name": "Формат",
    "type_code": "Код типа",
    "type_name": "Тип",
    "type_confidence": "Уверенность",
    "type_reason": "Обоснование типа",
    "type_evidence": "Индикаторы",
    "combo_code": "Код комбо",
    "combo_name": "Комбинация",
    "image_ocr_text": "OCR изображения",
    "image_description": "Описание изображения",
    "video_frame_ocr_text": "OCR кадров видео",
    "video_description": "Описание видео",
    "audio_transcript": "Транскрипция аудио",
    "voice_transcript": "Транскрипция голосового",
    "analysis_basis": "Основа классификации",
    "llm_used": "LLM использован",
    "manual_review_required": "Требует проверки",
    "classification_note": "Примечание",
    "views": "Просмотры",
    "forwards": "Пересылки",
    "replies": "Комментарии",
    "reactions": "Реакции",
    "random_seed": "Seed",
    "sample_order": "№ в выборке",
}


def _safe_str(val) -> str:
    """Безопасно конвертирует значение в строку, защищая от формул Excel."""
    s = str(val)
    # Если строка начинается с символов формулы — добавляем апостроф-префикс
    # Excel воспринимает = + - @ как начало формулы
    if s and s[0] in ("=", "+", "-", "@", "|"):
        return "'" + s
    return s


def _bool_to_str(val) -> str:
    if isinstance(val, bool):
        return "Да" if val else "Нет"
    if isinstance(val, str):
        if val.lower() in ("true", "да", "1", "yes"):
            return "Да"
        if val.lower() in ("false", "нет", "0", "no", ""):
            return "Нет"
    if isinstance(val, (int, float)):
        return "Да" if val else "Нет"
    return str(val) if val else "Нет"


def _prepare_post_row(post: dict) -> list:
    row = []
    for field in POST_COLUMNS:
        val = post.get(field)
        if hasattr(val, "strftime"):
            val = val.strftime("%Y-%m-%d %H:%M:%S")
        elif field in ("manual_review_required", "llm_used", "has_text", "has_photo",
                       "has_video", "has_video_note", "has_voice", "has_audio",
                       "has_poll", "has_gif", "has_sticker", "has_file"):
            val = _bool_to_str(val)
        elif isinstance(val, bool):
            val = "Да" if val else "Нет"
        elif isinstance(val, int) and field in ("manual_review_required", "llm_used"):
            val = "Да" if val else "Нет"
        elif isinstance(val, list):
            val = ", ".join(str(v) for v in val)
        elif val is None:
            val = ""
        elif isinstance(val, str):
            # Защита от строк которые Excel воспринимает как формулы
            if val and val[0] in ("=", "+", "-", "@", "|"):
                val = "'" + val
        row.append(val)
    return row


# ---------------------------------------------------------------------------
# Основная функция
# ---------------------------------------------------------------------------

def build_report(
    posts: list[dict],
    output_path: str,
    method_meta: dict | None = None,
) -> None:
    """
    Формирует Excel-отчёт с 14 листами.

    Args:
        posts: список постов с полной классификацией.
        output_path: путь для сохранения .xlsx.
        method_meta: словарь с методическими параметрами (период, seed и т.д.).
    """
    if not posts:
        logger.warning("Нет постов для отчёта")
        return

    analytics = Analytics(posts)

    xlsx_path = Path(output_path)
    if xlsx_path.suffix.lower() != ".xlsx":
        xlsx_path = xlsx_path.with_suffix(".xlsx")

    wb = openpyxl.Workbook()
    # Устанавливаем свойства книги чтобы Excel не выдавал предупреждение о восстановлении
    try:
        from openpyxl.packaging.core import DocumentProperties
        wb.properties = DocumentProperties()
        wb.properties.creator = "Telegram Post Analyzer"
        wb.properties.lastModifiedBy = "Telegram Post Analyzer"
    except Exception:
        pass

    # ---- Лист 1: posts_raw ----
    ws1 = wb.active
    ws1.title = "posts_raw"
    headers_ru = [POST_HEADERS_RU.get(c, c) for c in POST_COLUMNS]
    ws1.append(headers_ru)
    for cell in ws1[1]:
        _hdr(cell)
    ws1.row_dimensions[1].height = 30
    for row_idx, post in enumerate(posts, start=2):
        ws1.append(_prepare_post_row(post))
        for cell in ws1[row_idx]:
            _data(cell, row_idx)
    ws1.freeze_panes = "A2"
    ws1.auto_filter.ref = ws1.dimensions
    _auto_width(ws1)

    # ---- Лист 2: format_stats_total ----
    ws2 = wb.create_sheet("format_stats_total")
    _write_table(ws2,
                 ["format_code", "format_name", "count", "percent_total"],
                 analytics.format_stats_total())

    # ---- Лист 3: type_stats_total ----
    ws3 = wb.create_sheet("type_stats_total")
    _write_table(ws3,
                 ["type_code", "type_name", "count", "percent_total"],
                 analytics.type_stats_total())

    # ---- Лист 4: combo_stats_total ----
    ws4 = wb.create_sheet("combo_stats_total")
    _write_table(ws4,
                 ["combo_code", "combo_name", "format_code", "format_name",
                  "type_code", "type_name", "count", "percent_total"],
                 analytics.combo_stats_total())

    # ---- Лист 5: format_stats_by_segment ----
    ws5 = wb.create_sheet("format_stats_by_segment")
    _write_table(ws5,
                 ["segment", "segment_name", "format_code", "format_name",
                  "count", "percent_within_segment"],
                 analytics.format_stats_by_segment())

    # ---- Лист 6: type_stats_by_segment ----
    ws6 = wb.create_sheet("type_stats_by_segment")
    _write_table(ws6,
                 ["segment", "segment_name", "type_code", "type_name",
                  "count", "percent_within_segment"],
                 analytics.type_stats_by_segment())

    # ---- Лист 7: combo_stats_by_segment ----
    ws7 = wb.create_sheet("combo_stats_by_segment")
    _write_table(ws7,
                 ["segment", "segment_name", "combo_code", "combo_name",
                  "format_code", "format_name", "type_code", "type_name",
                  "count", "percent_within_segment"],
                 analytics.combo_stats_by_segment())

    # ---- Листы 8–10: pivot-таблицы ----
    _write_pivot(wb, "combo_pivot_total",   analytics.pivot_total())
    _write_pivot(wb, "combo_pivot_fashion", analytics.pivot_fashion(), title="Мода")
    _write_pivot(wb, "combo_pivot_beauty",  analytics.pivot_beauty(),  title="Красота")

    # ---- Лист 11: segment_comparison ----
    ws11 = wb.create_sheet("segment_comparison")
    _write_table(ws11,
                 ["combo_code", "combo_name",
                  "fashion_count", "fashion_percent",
                  "beauty_count", "beauty_percent",
                  "difference_percent_points", "intersection_status"],
                 analytics.segment_comparison())
    _color_comparison(ws11)

    # ---- Лист 12: stable_combinations ----
    ws12 = wb.create_sheet("stable_combinations")
    _write_table(ws12,
                 ["segment", "segment_name", "combo_code", "combo_name",
                  "count", "percent_within_segment", "rank_within_segment"],
                 analytics.stable_combinations())

    # ---- Лист 13: rare_combinations ----
    ws13 = wb.create_sheet("rare_combinations")
    _write_table(ws13,
                 ["segment", "segment_name", "combo_code", "combo_name",
                  "count", "percent_within_segment", "rank_from_bottom"],
                 analytics.rare_combinations())

    # ---- Лист 14: method_notes + summary_text ----
    ws14 = wb.create_sheet("method_notes")
    _write_method_notes(ws14, analytics, method_meta or {})

    wb.save(str(xlsx_path))
    logger.info(f"Отчёт сохранён: {xlsx_path} (14 листов, {len(posts)} постов)")

    # CSV-файлы
    _save_csv(posts, xlsx_path.with_name("posts_raw.csv"))
    _save_csv(analytics.combo_stats_total(),      xlsx_path.with_name("combo_stats_total.csv"))
    _save_csv(analytics.combo_stats_by_segment(), xlsx_path.with_name("combo_stats_by_segment.csv"))
    _save_csv(analytics.segment_comparison(),     xlsx_path.with_name("segment_comparison.csv"))


# ---------------------------------------------------------------------------
# Pivot-таблица
# ---------------------------------------------------------------------------

def _write_pivot(wb, sheet_name: str, pivot: dict, title: str = ""):
    ws = wb.create_sheet(sheet_name)
    formats = pivot["formats"]
    types   = pivot["types"]
    counts  = pivot["counts"]
    fn      = pivot["format_names"]
    tn      = pivot["type_names"]
    total   = pivot["total"]

    label = f"Сводная таблица форматы × типы{' (' + title + ')' if title else ''}"

    # Заголовки
    header = ["Формат \\ Тип"] + [f"{tc}\n{tn.get(tc,'')}" for tc in types] + ["Итого"]
    ws.append(header)
    for cell in ws[1]:
        _hdr(cell)
    ws.row_dimensions[1].height = 45

    # Данные
    yellow_fill = PatternFill(start_color=YELLOW, end_color=YELLOW, fill_type="solid")
    for row_idx, fc in enumerate(formats, start=2):
        row_label = f"{fc} — {fn.get(fc,'')}"
        row_vals = [counts[fc].get(tc, 0) for tc in types]
        row_total = sum(row_vals)
        ws.append([row_label] + row_vals + [row_total])

        for cell in ws[row_idx]:
            _data(cell, row_idx)
        # Итоговая колонка
        ws.cell(row=row_idx, column=len(types) + 2).fill = yellow_fill

    # Итоговая строка
    totals = ["Итого"]
    for tc in types:
        totals.append(sum(counts[fc].get(tc, 0) for fc in formats))
    totals.append(total)
    ws.append(totals)
    last = ws.max_row
    for cell in ws[last]:
        cell.fill = PatternFill(start_color=GREEN, end_color=GREEN, fill_type="solid")
        cell.font = Font(bold=True)
        cell.border = _thin()

    ws.freeze_panes = "B2"
    _auto_width(ws)


# ---------------------------------------------------------------------------
# Раскраска сравнительной таблицы
# ---------------------------------------------------------------------------

def _color_comparison(ws):
    """Раскрашивает строки segment_comparison по intersection_status."""
    status_col = 8  # intersection_status — 8-я колонка
    for row in ws.iter_rows(min_row=2):
        status = row[status_col - 1].value
        if status == "both_segments":
            color = "E8F5E9"  # зеленоватый
        elif status == "fashion_only":
            color = "E3F2FD"  # голубоватый
        elif status == "beauty_only":
            color = "FCE4EC"  # розоватый
        else:
            color = "FFFFFF"
        fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
        for cell in row:
            if cell.value is not None:
                cell.fill = fill


# ---------------------------------------------------------------------------
# Методические заметки
# ---------------------------------------------------------------------------

def _write_method_notes(ws, analytics: Analytics, meta: dict):
    ws.title = "method_notes"

    # Сводные выводы
    summary = analytics.summary_text(meta)
    ws.append(["Показатель", "Значение"])
    for cell in ws[1]:
        _hdr(cell)
    ws.row_dimensions[1].height = 25

    for row_idx, row in enumerate(summary, start=2):
        ws.append([row.get("Показатель", ""), row.get("Значение", "")])
        for cell in ws[row_idx]:
            _data(cell, row_idx)

    _auto_width(ws)
    ws.column_dimensions["A"].width = 45
    ws.column_dimensions["B"].width = 70


# ---------------------------------------------------------------------------
# Сохранение CSV
# ---------------------------------------------------------------------------

def _save_csv(rows: list[dict], path: Path):
    if not rows:
        return
    try:
        keys = list(rows[0].keys())
        with open(str(path), "w", newline="", encoding="utf-8-sig") as f:
            writer = csv.DictWriter(f, fieldnames=keys)
            writer.writeheader()
            for row in rows:
                # Конвертируем datetime и списки
                clean = {}
                for k, v in row.items():
                    if hasattr(v, "strftime"):
                        v = v.strftime("%Y-%m-%d %H:%M:%S")
                    if isinstance(v, list):
                        v = ", ".join(str(i) for i in v)
                    if isinstance(v, bool):
                        v = "Да" if v else "Нет"
                    clean[k] = v
                writer.writerow(clean)
        logger.info(f"CSV сохранён: {path}")
    except Exception as e:
        logger.error(f"Ошибка сохранения CSV {path}: {e}")


# ---------------------------------------------------------------------------
# Файл ручной проверки
# ---------------------------------------------------------------------------

MANUAL_REVIEW_COLUMNS = [
    "channel_url", "post_url", "post_date", "segment", "segment_name",
    "text", "format_code", "format_name",
    "predicted_type_code", "predicted_type_name",
    "type_confidence", "type_reason",
    "suggested_manual_type_code", "manual_comment",
]


def build_manual_review(posts: list[dict], output_path: str) -> None:
    """Создаёт файл для ручной проверки постов с низкой уверенностью."""
    review_posts = [p for p in posts if p.get("manual_review_required")]
    if not review_posts:
        logger.info("Нет постов, требующих ручной проверки")
        return

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "manual_review"

    headers = [
        "URL канала", "Ссылка на пост", "Дата", "Сегмент (EN)", "Сегмент",
        "Текст поста", "Код формата", "Формат",
        "Предсказанный код типа", "Предсказанный тип",
        "Уверенность", "Обоснование",
        "Исправленный код типа", "Комментарий",
    ]
    ws.append(headers)
    for cell in ws[1]:
        _hdr(cell)
    ws.row_dimensions[1].height = 30

    for row_idx, post in enumerate(review_posts, start=2):
        row = [
            post.get("channel_url", ""),
            post.get("post_url", ""),
            post.get("post_date", "").strftime("%Y-%m-%d") if hasattr(post.get("post_date"), "strftime") else post.get("post_date", ""),
            post.get("segment", ""),
            post.get("segment_name", ""),
            post.get("text", ""),
            post.get("format_code", ""),
            post.get("format_name", ""),
            post.get("type_code", ""),
            post.get("type_name", ""),
            post.get("type_confidence", ""),
            post.get("type_reason", ""),
            "",  # suggested_manual_type_code — заполняет исследователь
            "",  # manual_comment — заполняет исследователь
        ]
        ws.append(row)
        for cell in ws[row_idx]:
            _data(cell, row_idx)

    # Выделяем колонки для ручного заполнения
    yellow_fill = PatternFill(start_color=YELLOW, end_color=YELLOW, fill_type="solid")
    for row in ws.iter_rows(min_row=2, min_col=13, max_col=14):
        for cell in row:
            cell.fill = yellow_fill

    _auto_width(ws)
    ws.freeze_panes = "A2"

    path = Path(output_path)
    wb.save(str(path))
    logger.info(f"Файл ручной проверки: {path} ({len(review_posts)} постов)")


# ---------------------------------------------------------------------------
# Пересчёт аналитики после ручной проверки
# ---------------------------------------------------------------------------

def recalculate_with_manual(
    classified_path: str,
    manual_review_path: str,
    output_path: str,
) -> None:
    """
    Загружает заполненный файл ручной проверки и пересчитывает аналитику.
    Исправленные типы из manual_review заменяют предсказанные.
    """
    logger.info("Загрузка классифицированных постов...")
    wb_main = openpyxl.load_workbook(classified_path)
    ws_main = wb_main["posts_raw"]

    rows_main = list(ws_main.iter_rows(values_only=True))
    header = rows_main[0]
    posts = [dict(zip(header, row)) for row in rows_main[1:]]

    # Индексируем по post_url
    post_index = {p.get("Ссылка на пост"): p for p in posts}

    logger.info("Загрузка результатов ручной проверки...")
    wb_review = openpyxl.load_workbook(manual_review_path)
    ws_review = wb_review["manual_review"]
    rows_review = list(ws_review.iter_rows(values_only=True))
    review_header = rows_review[0]
    review_rows = [dict(zip(review_header, r)) for r in rows_review[1:]]

    corrected = 0
    accepted = 0
    from type_classifier import TYPE_NAMES

    for rev in review_rows:
        post_url = rev.get("Ссылка на пост")
        if not post_url:
            continue

        manual_code   = rev.get("Исправленный код типа")
        predicted_code = rev.get("Предсказанный код типа")

        if post_url not in post_index:
            continue

        # Если исправленный код заполнен — берём его
        if manual_code and str(manual_code).strip():
            try:
                code = int(str(manual_code).strip())
                post_index[post_url]["Код типа"] = code
                post_index[post_url]["Тип"] = TYPE_NAMES.get(code, str(code))
                post_index[post_url]["Основа классификации"] = "manual"
                post_index[post_url]["Требует проверки"] = "Нет"
                # Обновляем обоснование на стандартное для нового типа
                from type_classifier import TYPE_REASONS
                post_index[post_url]["Обоснование типа"] = TYPE_REASONS.get(code, "Тип исправлен вручную исследователем.")
                corrected += 1
                continue
            except ValueError:
                logger.warning(f"Неверный код типа в ручной проверке: {manual_code}")

        # Если исправленный пустой — берём предсказанный (вы согласились с ним)
        if predicted_code and str(predicted_code).strip():
            try:
                code = int(str(predicted_code).strip())
                post_index[post_url]["Код типа"] = code
                post_index[post_url]["Тип"] = TYPE_NAMES.get(code, str(code))
                post_index[post_url]["Основа классификации"] = "auto_accepted"
                post_index[post_url]["Требует проверки"] = "Нет"
                accepted += 1
            except ValueError:
                pass

    logger.info(f"Исправлено вручную: {corrected}, принято как есть: {accepted}")

    # Нормализуем ключи обратно
    def _normalize(post_dict: dict) -> dict:
        """Переводим русские заголовки обратно в английские ключи."""
        reverse = {v: k for k, v in POST_HEADERS_RU.items()}
        return {reverse.get(k, k): v for k, v in post_dict.items()}

    normalized_posts = [_normalize(p) for p in posts]

    build_report(normalized_posts, output_path,
                 method_meta={"Пересчёт после ручной проверки": f"да ({corrected} постов исправлено)"})
    logger.info(f"Пересчитанный отчёт сохранён: {output_path}")
