"""
auto_fix.py — автоматически исправляет противоречия типов на основе кодификатора.

Правила:
1. Объявление победителей розыгрыша (тип 5) → тип 1 (Информационный)
2. Мем-формат «покинул группу» → тип 6 (Развлекательный)
3. Гайд/образовательный + ссылки на покупку доминируют (>40% текста) → тип 4
4. Анонсовый + ссылка на покупку прямо сейчас → тип 4
5. Имиджевый + явная ссылка на покупку → тип 4
6. Нормализует все кривые названия типов

Запуск:
    python auto_fix.py \
      --input result_final_v2_run20260506_1052.xlsx \
      --manual-review result_final_v2_run20260506_1052_manual_review.xlsx \
      --output result_final_v4.xlsx
"""

import argparse
import logging
import re
import sys
from datetime import datetime
from pathlib import Path

import openpyxl

logging.basicConfig(
    level=logging.INFO,
    format="[%(asctime)s] %(levelname)s — %(message)s",
    datefmt="%Y-%m-%d %H:%M:%S",
)
logger = logging.getLogger("auto_fix")

TYPE_NAMES = {
    1: "Информационный",
    2: "Образовательный",
    3: "Анонсовый",
    4: "Продающий",
    5: "Вовлекающий",
    6: "Развлекательный",
    7: "Имиджево-вдохновляющий",
}

TYPE_REASONS = {
    1: "Пост передаёт актуальные сведения или новости без побуждения к немедленному действию.",
    2: "Пост обучает аудиторию — содержит инструкцию, совет, разбор или пошаговое объяснение.",
    3: "Пост анонсирует предстоящее событие — создаёт ожидание и подталкивает к планированию.",
    4: "Пост стимулирует покупку — содержит явное коммерческое предложение, цену или призыв к заказу.",
    5: "Пост вовлекает аудиторию в диалог — содержит вопрос, опрос, конкурс или призыв к реакции.",
    6: "Пост создаёт лёгкий эмоциональный фон — юмор, мем или ситуативный контент без утилитарной ценности.",
    7: "Пост транслирует философию и ценности бренда — эстетика, вдохновение, backstage без продающего посыла.",
}

# Все варианты кривых названий → правильный код
NAME_TO_CODE = {
    "информационный": 1, "информационное": 1, "информация": 1, "информационный тип": 1,
    "образовательный": 2, "образовательное": 2, "образование": 2, "обучающий": 2, "образователь": 2,
    "анонсовый": 3, "анонсовое": 3, "анонс": 3, "анонсирующий": 3, "анонсированный": 3, "анонсовой": 3, "анонсирован": 3,
    "продающий": 4, "продающее": 4, "продажный": 4, "продажное": 4, "продажа": 4,
    "продавающий": 4, "продавая": 4, "продвигающий": 4, "продаючий": 4,
    "продаоучий": 4, "продавщий": 4, "продаващий": 4, "продавющий": 4,
    "продавлющий": 4, "продавлюющий": 4, "продаучий": 4, "продаочий": 4,
    "вовлекающий": 5, "вовлекающее": 5, "вовлечение": 5,
    "развлекательный": 6, "развлекательное": 6, "развлечение": 6,
    "разлукательный": 6, "разлюкательный": 6, "развлечьельный": 6,
    "имиджево-вдохновляющий": 7, "имиджево вдохновляющий": 7, "имиджевый": 7,
    "вдохновляющий": 7, "имидже-вдохновляющий": 7, "имидувально-вдохновляющий": 7,
    "не определено": None, "undefined": None,
}


def normalize_code(code, name) -> int | None:
    """Возвращает нормализованный код типа."""
    if code is not None:
        try:
            c = int(str(code).strip().split('.')[0])
            if c in TYPE_NAMES:
                return c
        except (ValueError, AttributeError):
            pass
    if name:
        n = str(name).strip().lower()
        if n in NAME_TO_CODE:
            return NAME_TO_CODE[n]
        # Ищем по подстроке
        for key, val in NAME_TO_CODE.items():
            if key in n or n in key:
                return val
    return None


def count_sell_signals(text: str) -> int:
    """Считает количество продающих сигналов в тексте."""
    patterns = [
        r'wildberries\.ru', r'ozon\.ru', r'goldapple\.ru',
        r'поdружка\.ru', r'limestore\.com', r'faberlic\.com',
        r'befree\.ru', r'colins\.ru', r'zarina\.ru',
        r'\bwb\b', r'\bozon\b', r'золотое яблоко',
        r'купить', r'заказать', r'в корзину', r'промокод',
        r'скидка\s*\d+', r'\d+\s*[₽%]', r'₽\s*\d+',
        r'по ссылке', r'ссылка в шапке', r'только сегодня',
        r'уже на wb', r'уже на ozon',
    ]
    count = sum(1 for p in patterns if re.search(p, text.lower()))
    return count


def count_edu_signals(text: str) -> int:
    """Считает количество образовательных сигналов в тексте."""
    patterns = [
        r'как правильно', r'пошаговая', r'инструкция', r'лайфхак',
        r'совет', r'шаг \d', r'во-первых', r'во-вторых',
        r'в составе', r'как использовать', r'как работает',
        r'разберёмся', r'рассказываем', r'объясняем',
        r'гайд', r'рутина', r'нужно ли', r'можно ли',
        r'\d+\s*(совет|правил|шаг|лайфхак)',
    ]
    count = sum(1 for p in patterns if re.search(p, text.lower()))
    return count


def auto_fix_type(text: str, current_code: int, url: str) -> tuple[int | None, str | None]:
    """
    Проверяет пост по правилам кодификатора.
    Возвращает (новый_код, причина) или (None, None) если исправление не нужно.
    """
    text_lower = text.lower()

    # Правило 1: объявление победителей → информационный
    if current_code == 5:
        winner_signals = ['победитель', 'победител', '🏆', 'результаты розыгрыша',
                          'результаты конкурса', 'итоги конкурса', 'итоги розыгрыша',
                          'счастливчик', 'победили', 'выиграл']
        active_signals = ['участвуй', 'конкурс!', 'розыгрыш!', 'условия участия',
                          'нажми', 'подпишись', 'отметь']
        has_winners = any(w in text_lower for w in winner_signals)
        has_active = any(w in text_lower for w in active_signals)
        if has_winners and not has_active:
            return 1, "Объявление победителей розыгрыша — информационный, не вовлекающий."

    # Правило 2: мем-формат → развлекательный
    meme_signals = ['покинул(а) группу', 'покинул группу', 'покинула группу',
                    'вышел из чата', 'вышла из чата', 'добавился в группу']
    if current_code != 6 and any(m in text_lower for m in meme_signals):
        return 6, "Мем-формат «покинул группу» — развлекательный тип."

    # Правило 3: образовательный + ссылки доминируют → продающий
    if current_code == 2:
        sell = count_sell_signals(text)
        edu = count_edu_signals(text)
        # Считаем количество ссылок в тексте
        links = len(re.findall(r'https?://\S+|\[.+?\]\(.+?\)', text))
        if sell >= 3 or links >= 4:
            return 4, f"Образовательный контент с доминирующими ссылками на покупку ({links} ссылок, {sell} продающих сигналов) — продающий."

    # Правило 4: анонсовый без временного маркера будущего
    if current_code == 3:
        sell = count_sell_signals(text)
        links = len(re.findall(r'https?://\S+|\[.+?\]\(.+?\)', text))

        # Проверяем наличие временного маркера будущего
        future_markers = ['скоро', 'завтра', 'сохрани', 'не пропусти', 'прямой эфир',
                          'coming soon', 'save the date', 'через', 'предстоит',
                          'в январ', 'в феврал', 'в март', 'в апрел', 'в мае', 'в июн',
                          'в июл', 'в август', 'в сентябр', 'в октябр', 'в ноябр', 'в декабр',
                          'в эту пятницу', 'в эту субботу', 'на этой неделе',
                          'на следующей неделе', 'готовится', 'выйдет', 'появится',
                          'запустим', 'откроется', 'ждите', 'следите']
        has_future = any(m in text_lower for m in future_markers)
        # Дата в тексте
        has_date = bool(re.search(r'\d{1,2}\s*(января|февраля|марта|апреля|мая|июня|июля|августа|сентября|октября|ноября|декабря)|\d{1,2}[.:]\d{2}', text_lower))

        # Продажа прямо сейчас
        now_signals = ['уже на wb', 'уже на ozon', 'уже в продаже', 'уже доступно',
                       'купить сейчас', 'заказать сейчас', 'ссылка в шапке',
                       'уже открыт', 'уже открылся', 'уже появились', 'уже доступна']
        has_now = any(s in text_lower for s in now_signals)

        if has_now and sell >= 2:
            return 4, "Анонс с активной продажей прямо сейчас — продающий."

        # Нет временного маркера будущего + есть ссылки → продающий
        if not has_future and not has_date and links >= 1 and sell >= 1:
            return 4, "Отнесён к анонсовому, но нет временного маркера будущего — пост о текущих новинках со ссылками на покупку, продающий."

        # Нет маркера будущего + нет ссылок → информационный
        if not has_future and not has_date and links == 0 and len(text) > 20:
            return 1, "Отнесён к анонсовому, но нет временного маркера будущего — пост сообщает о факте, информационный."

    # Правило 5: имиджевый + явная ссылка на покупку → продающий
    if current_code == 7:
        sell = count_sell_signals(text)
        links = len(re.findall(r'https?://\S+|\[.+?\]\(.+?\)', text))
        if sell >= 2 and links >= 2:
            return 4, f"Имиджевый пост с явными ссылками на покупку ({links} ссылок) — продающий."

    return None, None


def parse_args():
    parser = argparse.ArgumentParser()
    parser.add_argument("--input",         required=True)
    parser.add_argument("--manual-review", default="", dest="manual_review")
    parser.add_argument("--output",        default="result_auto_fixed.xlsx")
    return parser.parse_args()


def main():
    args = parse_args()
    logger.info("=" * 60)
    logger.info("Автоматическое исправление противоречий типов")
    logger.info("=" * 60)

    if not Path(args.input).exists():
        logger.error(f"Файл не найден: {args.input}")
        sys.exit(1)

    wb = openpyxl.load_workbook(args.input)
    ws = wb["posts_raw"] if "posts_raw" in wb.sheetnames else wb.active

    rows = list(ws.iter_rows(values_only=False))
    headers = [str(cell.value).strip() if cell.value else f"col_{i}"
               for i, cell in enumerate(rows[0])]

    def col(*names):
        for n in names:
            if n in headers:
                return headers.index(n)
        return None

    url_idx     = col("Ссылка на пост")
    type_idx    = col("Тип", "Название типа")
    code_idx    = col("Код типа")
    reason_idx  = col("Обоснование типа", "Обоснование")
    combo_idx   = col("Комбинация")
    combo_c_idx = col("Код комбо")
    format_idx  = col("Формат", "Название формата")
    format_c_idx = col("Код формата")
    basis_idx   = col("Основа классификации")
    review_idx  = col("Требует проверки")
    text_idx    = col("Текст поста")

    # Загружаем ручные исправления
    manual_fixes = {}
    if args.manual_review and Path(args.manual_review).exists():
        wb_mr = openpyxl.load_workbook(args.manual_review)
        ws_mr = wb_mr.active
        mr_rows = list(ws_mr.iter_rows(values_only=True))
        mr_h = [str(h) for h in mr_rows[0]]
        mr_url = mr_h.index("Ссылка на пост") if "Ссылка на пост" in mr_h else None
        mr_fix = mr_h.index("Исправленный код типа") if "Исправленный код типа" in mr_h else None
        if mr_url is not None and mr_fix is not None:
            for row in mr_rows[1:]:
                u = str(row[mr_url] or "").strip()
                f = row[mr_fix]
                if u and f and str(f).strip():
                    try:
                        manual_fixes[u] = int(str(f).strip())
                    except ValueError:
                        pass
        logger.info(f"Ручных исправлений: {len(manual_fixes)}")

    stats = {
        "normalized": 0,
        "manual_applied": 0,
        "auto_fixed": 0,
        "undefined_kept": 0,
    }
    fix_log = []

    for row in rows[1:]:
        vals = {h: row[i].value for i, h in enumerate(headers) if i < len(row)}

        url         = str(vals.get("Ссылка на пост", "") or "").strip()
        raw_code    = vals.get("Код типа")
        raw_name    = str(vals.get("Тип", "") or vals.get("Название типа", "") or "")
        text        = str(vals.get("Текст поста", "") or "")
        format_name = str(vals.get("Формат", "") or vals.get("Название формата", "") or "")
        format_code = vals.get("Код формата")

        # 1. Ручные исправления (приоритет)
        url_clean = url.rstrip("/")
        matched = next((k for k in manual_fixes if k.rstrip("/") == url_clean), None)
        if matched:
            new_code = manual_fixes[matched]
            if code_idx is not None: row[code_idx].value = new_code
            if type_idx is not None: row[type_idx].value = TYPE_NAMES[new_code]
            if reason_idx is not None: row[reason_idx].value = "Тип исправлен вручную исследователем."
            if basis_idx is not None: row[basis_idx].value = "manual"
            if review_idx is not None: row[review_idx].value = "Нет"
            current_code = new_code
            stats["manual_applied"] += 1
        else:
            # 2. Нормализация названия
            current_code = normalize_code(raw_code, raw_name)
            if current_code != raw_code or TYPE_NAMES.get(current_code, "") != raw_name:
                if code_idx is not None: row[code_idx].value = current_code
                if type_idx is not None: row[type_idx].value = TYPE_NAMES.get(current_code, raw_name) if current_code else raw_name
                stats["normalized"] += 1

            # 3. Автоисправление противоречий
            if current_code and text.strip():
                new_code, reason = auto_fix_type(text, current_code, url)
                if new_code and new_code != current_code:
                    if code_idx is not None: row[code_idx].value = new_code
                    if type_idx is not None: row[type_idx].value = TYPE_NAMES[new_code]
                    if reason_idx is not None: row[reason_idx].value = reason
                    if basis_idx is not None: row[basis_idx].value = "auto_fixed"
                    if review_idx is not None: row[review_idx].value = "Нет"
                    current_code = new_code
                    stats["auto_fixed"] += 1
                    fix_log.append(f"{url} | {TYPE_NAMES.get(normalize_code(raw_code, raw_name), raw_name)} → {TYPE_NAMES[new_code]} | {reason}")

            if current_code is None:
                stats["undefined_kept"] += 1

        # 4. Обновляем обоснование если пустое или кривое
        current_reason = str(vals.get("Обоснование типа", "") or vals.get("Обоснование", "") or "")
        if current_code and (len(current_reason) < 10 or not any(c in current_reason for c in 'аеиоуыэюяАЕИОУЫЭЮЯ')):
            if reason_idx is not None:
                row[reason_idx].value = TYPE_REASONS.get(current_code, "")

        # 5. Пересчитываем комбинацию
        if current_code and format_code:
            try:
                fc = int(str(format_code).strip())
                combo_code = f"{fc}.{current_code}"
                combo_name = f"{format_name} + {TYPE_NAMES.get(current_code, str(current_code))}"
                if combo_c_idx is not None: row[combo_c_idx].value = combo_code
                if combo_idx is not None: row[combo_idx].value = combo_name
            except (ValueError, TypeError):
                pass

    logger.info(f"Нормализовано названий:    {stats['normalized']}")
    logger.info(f"Применено ручных правок:   {stats['manual_applied']}")
    logger.info(f"Автоисправлено:            {stats['auto_fixed']}")
    logger.info(f"Осталось «Не определено»:  {stats['undefined_kept']}")

    if fix_log:
        logger.info("\nАвтоисправления:")
        for line in fix_log:
            logger.info(f"  {line}")

    # Сохраняем промежуточный файл
    p = Path(args.output)
    run_str = datetime.now().strftime("%Y%m%d_%H%M")
    output_path = str(p.parent / f"{p.stem}_run{run_str}{p.suffix}")
    wb.save(output_path)
    logger.info(f"Сохранено: {output_path}")

    # Пересчитываем аналитику
    logger.info("Пересчитываем аналитику...")
    try:
        import pandas as pd
        from report_builder import build_report, build_manual_review

        df = pd.read_excel(output_path, sheet_name="posts_raw")

        key_map = {
            "URL канала": "channel_url", "Username": "channel_username",
            "ID поста": "post_id", "Ссылка на пост": "post_url",
            "Дата публикации": "post_date", "Месяц": "month",
            "Сегмент (EN)": "segment", "Сегмент": "segment_name",
            "Текст поста": "text", "Код формата": "format_code",
            "Формат": "format_name", "Код типа": "type_code",
            "Тип": "type_name", "Уверенность": "type_confidence",
            "Обоснование типа": "type_reason", "Код комбо": "combo_code",
            "Комбинация": "combo_name", "Основа классификации": "analysis_basis",
            "Требует проверки": "manual_review_required",
            "Есть текст": "has_text", "Фото": "has_photo", "Видео": "has_video",
            "Кружок": "has_video_note", "Голосовое": "has_voice",
            "Аудио": "has_audio", "Опрос": "has_poll", "GIF": "has_gif",
            "Стикер": "has_sticker", "Файл": "has_file",
            "Кол-во медиа": "media_count", "Просмотры": "views",
            "Пересылки": "forwards", "Комментарии": "replies",
            "Реакции": "reactions", "Seed": "random_seed", "№ в выборке": "sample_order",
            "OCR изображения": "image_ocr_text", "Описание изображения": "image_description",
            "OCR кадров видео": "video_frame_ocr_text", "Описание видео": "video_description",
            "Транскрипция аудио": "audio_transcript", "Транскрипция голосового": "voice_transcript",
            "LLM использован": "llm_used", "Примечание": "classification_note",
            "Индикаторы": "type_evidence",
        }

        posts = []
        for _, r in df.iterrows():
            post = {}
            for k, v in r.items():
                post[key_map.get(str(k), str(k))] = v
            posts.append(post)

        final_path = output_path.replace(".xlsx", "_final.xlsx")
        build_report(posts, final_path)

        # Manual review для оставшихся "Не определено"
        mr_posts = [p for p in posts if str(p.get("type_name", "")).strip() in
                    ("Не определено", "", "nan") or p.get("type_code") is None]
        if mr_posts:
            mr_path = output_path.replace(".xlsx", "_manual_review.xlsx")
            build_manual_review(mr_posts, mr_path)
            logger.info(f"На ручную проверку: {len(mr_posts)} постов → {mr_path}")

        logger.info(f"Финальный отчёт: {final_path}")
        print(f"\n✅ Готово!")
        print(f"   Финальный отчёт:        {final_path}")
        print(f"   Нормализовано:          {stats['normalized']}")
        print(f"   Ручных правок:          {stats['manual_applied']}")
        print(f"   Автоисправлено:         {stats['auto_fixed']}")
        print(f"   Осталось неопределено:  {stats['undefined_kept']}")

    except Exception as e:
        logger.error(f"Ошибка пересчёта аналитики: {e}")
        import traceback; traceback.print_exc()
        print(f"\n⚠️ Файл сохранён: {output_path}")
        print(f"   Запустите отдельно: python normalize.py --input {output_path}")


if __name__ == "__main__":
    main()
