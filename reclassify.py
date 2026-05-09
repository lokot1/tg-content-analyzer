"""
reclassify.py — быстрая переклассификация на основе кэшированных медиа-данных.

Логика:
1. Посты где basis="manual" — пропускаем (ваши ручные правки)
2. Посты где есть текст → qwen2.5:7b (лучше понимает русский)
3. Посты без текста с медиа → сначала llava описывает, потом qwen классифицирует
4. Посты где rules уверен (>= threshold) → оставляем как есть

Запуск:
    python reclassify.py \\
      --input result_final.xlsx \\
      --output result_v3.xlsx \\
      --vision-model llava:7b \\
      --text-model qwen2.5:7b \\
      --threshold 0.51
"""

import argparse
import logging
import os
import sys
from datetime import datetime
from pathlib import Path

import openpyxl

from utils import setup_logger
from type_classifier import classify_type, classify_by_rules, _classify_by_ollama, TYPE_NAMES
from report_builder import build_report, build_manual_review

SEGMENT_NAMES = {"fashion": "Мода", "beauty": "Красота"}

HEADER_MAP = {
    "URL канала": "channel_url",
    "Username канала": "channel_username",
    "Username": "channel_username",
    "ID поста": "post_id",
    "Ссылка на пост": "post_url",
    "Дата публикации": "post_date",
    "Месяц": "month",
    "Месяц (YYYY-MM)": "month",
    "Сегмент (EN)": "segment",
    "Сегмент": "segment_name",
    "Текст поста": "text",
    "Есть текст": "has_text",
    "Есть фото": "has_photo",
    "Есть видео": "has_video",
    "Есть кружок": "has_video_note",
    "Есть голосовое": "has_voice",
    "Есть аудио": "has_audio",
    "Есть опрос": "has_poll",
    "Опрос": "has_poll",
    "GIF": "has_gif",
    "Стикер": "has_sticker",
    "Файл": "has_file",
    "Кол-во медиа": "media_count",
    "Код формата": "format_code",
    "Формат": "format_name",
    "Название формата": "format_name",
    "Код типа": "type_code",
    "Тип": "type_name",
    "Название типа": "type_name",
    "Уверенность": "type_confidence",
    "Обоснование типа": "type_reason",
    "Индикаторы": "type_evidence",
    "Код комбо": "combo_code",
    "Комбинация": "combo_name",
    "OCR изображения": "image_ocr_text",
    "Описание изображения": "image_description",
    "OCR кадров видео": "video_frame_ocr_text",
    "Описание видео": "video_description",
    "Транскрипция аудио": "audio_transcript",
    "Транскрипция голосового": "voice_transcript",
    "Основа классификации": "analysis_basis",
    "LLM использован": "llm_used",
    "Требует проверки": "manual_review_required",
    "Примечание к классификации": "classification_note",
    "Примечание": "classification_note",
    "Просмотры": "views",
    "Пересылки": "forwards",
    "Комментарии": "replies",
    "Реакции": "reactions",
    "Seed выборки": "random_seed",
    "Seed": "random_seed",
    "№ в выборке": "sample_order",
    "Номер в выборке": "sample_order",
}


def parse_args():
    parser = argparse.ArgumentParser(description="Переклассификация типов")
    parser.add_argument("--input",        required=True)
    parser.add_argument("--output",       default="result_reclassified.xlsx")
    parser.add_argument("--log",          default="logs_reclassify.txt")
    parser.add_argument("--vision-model", default="llava:7b",    dest="vision_model")
    parser.add_argument("--text-model",   default="qwen2.5:7b",  dest="text_model")
    parser.add_argument("--threshold",    type=float, default=0.51)
    parser.add_argument("--mode",         default="hybrid", choices=["rules", "hybrid"])
    return parser.parse_args()


def load_posts(path: str) -> list[dict]:
    wb = openpyxl.load_workbook(path)
    ws = wb["posts_raw"] if "posts_raw" in wb.sheetnames else wb.active
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []
    headers = [str(h).strip() if h else f"col_{i}" for i, h in enumerate(rows[0])]
    posts = []
    for row in rows[1:]:
        raw = dict(zip(headers, row))
        if not any(raw.values()):
            continue
        post = {}
        for k, v in raw.items():
            key = HEADER_MAP.get(k, k)
            post[key] = v if v is not None else ""
        posts.append(post)
    return posts


def _bool(val) -> bool:
    if isinstance(val, bool):
        return val
    if isinstance(val, str):
        return val.strip().lower() in ("да", "true", "1", "yes")
    return bool(val)


def _str(val) -> str:
    return str(val).strip() if val else ""


def main():
    args = parse_args()
    logger = setup_logger(args.log)
    logger.info("=" * 60)
    logger.info("Переклассификация: qwen2.5:7b (текст) + llava:7b (медиа)")
    logger.info("=" * 60)
    logger.info(f"Входной файл: {args.input}")
    logger.info(f"Текстовая модель: {args.text_model}")
    logger.info(f"Vision модель:    {args.vision_model}")
    logger.info(f"Порог rules:      {args.threshold}")

    if not Path(args.input).exists():
        logger.error(f"Файл не найден: {args.input}")
        sys.exit(1)

    posts = load_posts(args.input)
    logger.info(f"Загружено постов: {len(posts)}")

    stats = {
        "skipped_manual": 0,
        "kept_rules": 0,
        "via_qwen": 0,
        "via_llava_qwen": 0,
        "manual_review": 0,
    }

    for i, post in enumerate(posts):
        basis = _str(post.get("analysis_basis", ""))

        # ── 1. Пропускаем ручные правки — но обновляем combo_name ──
        if basis == "manual":
            stats["skipped_manual"] += 1
            # Обновляем комбо если оно содержит "Не определено"
            combo = _str(post.get("combo_name", ""))
            if "Не определено" in combo or not combo:
                tc = post.get("type_code")
                fc = post.get("format_code")
                tn = post.get("type_name") or (TYPE_NAMES.get(int(tc), "") if tc else "")
                fn = _str(post.get("format_name", ""))
                post["combo_code"] = f"{fc if fc else '?'}.{tc if tc else '?'}"
                post["combo_name"] = f"{fn} + {tn}"
            continue

        text         = _str(post.get("text", ""))
        format_name  = _str(post.get("format_name", ""))
        extra_text   = " ".join(filter(None, [
            _str(post.get("image_ocr_text")),
            _str(post.get("video_frame_ocr_text")),
            _str(post.get("audio_transcript")),
            _str(post.get("voice_transcript")),
        ]))
        img_desc = " ".join(filter(None, [
            _str(post.get("image_description")),
            _str(post.get("video_description")),
        ]))
        has_poll    = _bool(post.get("has_poll"))
        has_gif     = _bool(post.get("has_gif"))
        has_sticker = _bool(post.get("has_sticker"))
        has_media   = any([
            _bool(post.get("has_photo")),
            _bool(post.get("has_video")),
            _bool(post.get("has_video_note")),
            _bool(post.get("has_voice")),
        ])

        # ── 2. Считаем длину контента ────────────────────────────────
        all_content = " ".join(filter(None, [text, extra_text, img_desc]))
        word_count = len(all_content.split())

        # ── 3. Очень короткий пост БЕЗ медиа-данных → сразу на проверку
        if word_count < 5 and not img_desc and not extra_text:
            # Нечего анализировать — честно отправляем на проверку
            from type_classifier import TypeResult
            type_result = TypeResult(
                type_code=None, type_name="Не определено",
                type_confidence=0.0,
                type_reason="Слишком мало контента для уверенной классификации.",
                type_evidence=[],
                analysis_basis="insufficient_content",
                llm_used=False, manual_review_required=True,
            )
            stats["manual_review"] += 1
            # Обновляем пост и идём дальше
            post["type_code"]              = type_result.type_code
            post["type_name"]              = type_result.type_name
            post["type_confidence"]        = type_result.type_confidence
            post["type_reason"]            = type_result.type_reason
            post["type_evidence"]          = type_result.type_evidence
            post["analysis_basis"]         = type_result.analysis_basis
            post["llm_used"]               = type_result.llm_used
            post["manual_review_required"] = type_result.manual_review_required
            fc = post.get("format_code")
            post["combo_code"] = f"{fc if fc else '?'}.?"
            post["combo_name"] = f"{format_name} + Не определено"
            seg = _str(post.get("segment", ""))
            if not seg:
                sn = _str(post.get("segment_name", ""))
                seg = "fashion" if sn in ("Мода", "fashion") else "beauty"
            post["segment"]      = seg
            post["segment_name"] = SEGMENT_NAMES.get(seg, seg)
            continue

        # ── 4. LLM-first: qwen читает пост + словари как подсказка ──────
        if args.mode == "hybrid":
            # Получаем подсказку от словарей
            rules_result = classify_by_rules(text, has_poll, has_gif, has_sticker, extra_text)
            hint = ""
            if rules_result.type_code and rules_result.type_confidence >= 0.4:
                hint = f"\n\nHint from keyword analysis: possibly type {rules_result.type_code} ({rules_result.type_name}), confidence {rules_result.type_confidence:.2f}. Consider this but rely on your own analysis."

            type_result = _classify_by_ollama(
                text=text + hint,
                format_name=format_name,
                extra_text=extra_text,
                image_description=img_desc,
                model=args.text_model,
            )
            stats["via_qwen"] += 1

            # Если qwen не уверен на коротком посте — используем словари
            if type_result.type_code is None or type_result.type_confidence < 0.45:
                if rules_result.type_code and rules_result.type_confidence >= args.threshold:
                    type_result = rules_result
                    if img_desc:
                        type_result.analysis_basis = "text+local_vision"
                    stats["kept_rules"] += 1
                    stats["via_qwen"] -= 1
                else:
                    type_result.manual_review_required = True

        else:
            type_result = classify_by_rules(text, has_poll, has_gif, has_sticker, extra_text)
            stats["kept_rules"] += 1

        # Финальный fallback
        if type_result.type_code is None:
            type_result.analysis_basis = "hybrid(rules_fallback)"
            type_result.manual_review_required = True

        # Обновляем пост
        post["type_code"]              = type_result.type_code
        post["type_name"]              = type_result.type_name
        post["type_confidence"]        = type_result.type_confidence
        post["type_reason"]            = type_result.type_reason
        post["type_evidence"]          = type_result.type_evidence
        post["analysis_basis"]         = type_result.analysis_basis
        post["llm_used"]               = type_result.llm_used
        post["manual_review_required"] = type_result.manual_review_required

        fc = post.get("format_code")
        tc = type_result.type_code
        post["combo_code"] = f"{fc if fc else '?'}.{tc if tc else '?'}"
        post["combo_name"] = f"{format_name} + {type_result.type_name}"

        seg = _str(post.get("segment", ""))
        if not seg:
            sn = _str(post.get("segment_name", ""))
            seg = "fashion" if sn in ("Мода", "fashion") else "beauty"
        post["segment"]      = seg
        post["segment_name"] = SEGMENT_NAMES.get(seg, seg)

        if type_result.manual_review_required:
            stats["manual_review"] += 1

        if (i + 1) % 100 == 0:
            logger.info(
                f"  {i+1}/{len(posts)} — "
                f"rules: {stats['kept_rules']}, "
                f"qwen: {stats['via_qwen']}, "
                f"manual: {stats['skipped_manual']}"
            )

    logger.info("=" * 60)
    logger.info(f"Пропущено (ручные правки): {stats['skipped_manual']}")
    logger.info(f"Оставлено rules:           {stats['kept_rules']}")
    logger.info(f"Через qwen:                {stats['via_qwen']}")
    logger.info(f"Через llava+qwen:          {stats['via_llava_qwen']}")
    logger.info(f"На ручную проверку:        {stats['manual_review']}")
    logger.info("=" * 60)

    # Автоимя файла
    p = Path(args.output)
    run_str = datetime.now().strftime("%Y%m%d_%H%M")
    output_path = str(p.parent / f"{p.stem}_run{run_str}{p.suffix}")

    build_report(posts, output_path)
    manual_path = Path(output_path).with_name(
        Path(output_path).stem + "_manual_review.xlsx"
    )
    build_manual_review(posts, str(manual_path))

    print(f"\n✅ Готово!")
    print(f"   Отчёт:           {output_path}")
    print(f"   Ручная проверка: {manual_path}")
    print(f"\n   Пропущено (ручные правки): {stats['skipped_manual']}")
    print(f"   Через qwen:                {stats['via_qwen']}")
    print(f"   На ручную проверку:        {stats['manual_review']}")


if __name__ == "__main__":
    main()
